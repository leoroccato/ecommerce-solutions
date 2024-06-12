from datetime import datetime, timedelta
import pandas as pd
from barcodeReader import *
from dataExtracter import *
from openpyxl import Workbook
import os


wb = Workbook()
ws = wb.active

r = 1
ws.cell(row=1, column=1).value = "Filename"
ws.cell(row=1, column=2).value = "Linha Digitável"
ws.cell(row=1, column=3).value = "Data de Vencimento"
ws.cell(row=1, column=4).value = "Valor"

pasta_boletos = r'C:\Users\leoro\Desktop\Projetos\DMStores\Automatização de Devoluções\barcodeReader\boletos'

print("Iniciando extração dos códigos de barra...")
print()

for arquivo in os.listdir(pasta_boletos):
    if arquivo.endswith(".pdf"):
        pdf_path = os.path.join(pasta_boletos, arquivo)

        try: 
            barcode = BarcodeReader(pdf_path)
            linha_dig = linha_digitavel(barcode)
            data_venc = data_extracter_venc(pdf_path)
            valor = data_extracter_valor(pdf_path)
        except Exception as e:
            barcode = False
            print(f"Tive algum erro ao ler o código de barras {e}")
        else:
            if not barcode: 
                print(f"Não consegui obter código de barras para {pdf_path}")
            else:
                print(f"Sucesso: {arquivo}")
                r += 1

                # Grava no Excel
                ws.cell(row=r, column=1).value = arquivo
                ws.cell(row=r, column=2).value = linha_dig
                ws.cell(row=r, column=3).value = data_venc
                ws.cell(row=r, column=4).value = valor

wb.save("Taxas_DMStores.xlsx")

