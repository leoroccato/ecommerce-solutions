from report import generate_report
from datetime import datetime
import pyautogui
import time
import openpyxl
import mss.tools
import os

print("Iniciando execução, certifique-se que o emulador esta aberto e no App Mercado Pago... ")
print()

# Pre-processing [Caminhos, criação das listas]
path_excel = r'C:\Users\leoro\Desktop\Projetos\DMStores\Automatização de Devoluções\barcodeReader\Taxas_DMStores.xlsx'
comprovantes = r'C:\Users\leoro\Desktop\Projetos\DMStores\Automatização de Devoluções\barcodeReader\comprovantes'

hoje = datetime.now().day

workbook = openpyxl.load_workbook(path_excel)
sheet = workbook.active
valores_coluna_b = []
valores_coluna_a = []
valores_coluna_c = []
valores_coluna_d = []

lista_pedidos_pagos = []
lista_valor_pagos =[]

# Iteração sobre a tabela do Excel
# Captura códigos de barra e guarda na lista
for linha in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
    for celula in linha:
        # Add valores das celulas a lista
        codigo = str(celula.value)
        codigo_formatado = codigo.replace(".", "").replace(" ", "")
        valores_coluna_b.append(codigo_formatado)
# Captura números de Pedido e guarda na lista
for linha in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for celula in linha:
        # Add valores das celulas a lista
        filename = str(celula.value)
        pedido = filename[:5]
        valores_coluna_a.append(pedido)
# Captura dia de vencimento e guarda na lista
for linha in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3):
    for celula in linha:
        # Add valores das celulas a lista
        data_venc = str(celula.value)
        dia_venc = data_venc[:2]
        valores_coluna_c.append(dia_venc)
# Captura valores dos boletos e guarda na lista
for linha in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
    for celula in linha:
        # Add valores das celulas a lista
        valor = str(celula.value)
        valores_coluna_d.append(valor)

# Loop de automação no App Mercado Pago
for code, name, dia_venc, valor in zip(valores_coluna_b, valores_coluna_a, valores_coluna_c, valores_coluna_d):
    time.sleep(5)
    # Clica em Contas e Serviços
    pyautogui.click(3154, 829)
    time.sleep(5)
    # Clica em Pagar Nova Conta
    pyautogui.click(3304, 928)
    time.sleep(5)
    # Clica em Pagar com Código de Barras
    pyautogui.click(3339, 457)
    time.sleep(5)
    # Cola o Código do Boleto
    pyautogui.typewrite(code, interval=0.1)
    time.sleep(5)
    # Clica em Continuar
    pyautogui.click(3304, 928)
    time.sleep(5)
    # Obtenha a cor do pixel nas coordenadas especificadas
    cor_pixel = pyautogui.pixel(3386, 532)

    # Se a cor for igual a (255, 0, 0) - vermelho, faça algo
    if cor_pixel == (245, 245, 245):
        print(f"Boleto ja foi pago {name}")
        time.sleep(5)
        pyautogui.click(3317, 923)
    # Verificação se o boleto vence hoje
    else:
        if dia_venc == hoje:
            # Clica em Pagar Agora Down
            pyautogui.click(3312, 931)
            time.sleep(5)
        else:
            # Clica em Pagar Agora Up
            pyautogui.click(3310, 860)
            time.sleep(5)
        # Clica em Pagar
        pyautogui.click(3312, 931)
        time.sleep(5)
        # Inserir senha
        pyautogui.typewrite("1234", interval=0.2)
        time.sleep(5)
        # Aperta Enter
        pyautogui.press('enter')
        time.sleep(10)
        # Clica em Ver Comprovante
        pyautogui.click(3342, 541)
        time.sleep(5)
        # Rotina para captura de tela para comprovante
        with mss.mss() as sct:
            # Get information of monitor 2
            monitor_number = 1
            mon = sct.monitors[monitor_number]

            # The screen part to capture
            monitor = {
                "top": mon["top"],
                "left": mon["left"],
                "width": mon["width"],
                "height": mon["height"],
                "mon": monitor_number,
            }
            if not os.path.exists(comprovantes):
                os.makedirs(comprovantes)
            output = os.path.join(comprovantes, f"{name}.png".format(**monitor))

            # Grab the data
            sct_img = sct.grab(monitor)

            # Save to the picture file
            mss.tools.to_png(sct_img.rgb, sct_img.size, output=output)
        # Clica em Voltar
        pyautogui.click(3108, 239)
        time.sleep(8)
        # Clica no X
        pyautogui.click(3498, 238)
        time.sleep(10)
        print(f"Paguei o Pedido {name}, no valor de R$ {valor}")
        lista_pedidos_pagos.append(name)
        lista_valor_pagos.append(valor)

generate_report(lista_pedidos_pagos, lista_valor_pagos)













